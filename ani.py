import openpyxl

import os 
os.getcwd()
os.chdir('C:\\Users\\anich\\Desktop\\PROJECTS\\python_categorize_from_excel\\')
 
import pandas as pd 
file = 'data.xlsx'
data = pd.ExcelFile(file)

df = data.parse('Sheet1')
df.info


ps = openpyxl.load_workbook('data.xlsx')
sheet = ps['Sheet1']
sheet.max_row

xbox = ["Series S","Series X"]
playstation = ["PS5","PS4","PS3","PS2","PSP"]

for row in range(2, sheet.max_row + 1):
    homeSetup = sheet['C' + str(row)].value
    workSetup = sheet['D' + str(row)].value
 
    my_list = homeSetup.split(",")

    my_list2 = workSetup.split(",")

    xbox_count = 0
    playstation_count = 0


    if('Series S' in my_list):
        xbox_count += 1                          
    if('Series X' in my_list):
        xbox_count += 1

    if('Series S' in my_list2):
        xbox_count += 1                          
    if('Series X' in my_list2):
        xbox_count += 1 


    if('PS5' in my_list2):
        playstation_count += 1 
    if('PS4' in my_list2):
        playstation_count += 1 
    if('PS3' in my_list2):
        playstation_count += 1 
    if('PS2' in my_list2):
        playstation_count += 1 
    if('PSP' in my_list2):
        playstation_count += 1 

    if('PS5' in my_list):
        playstation_count += 1 
    if('PS4' in my_list):
        playstation_count += 1 
    if('PS3' in my_list):
        playstation_count += 1 
    if('PS2' in my_list):
        playstation_count += 1 
    if('PSP' in my_list):
        playstation_count += 1 

    print_for_xbox = "Value of xbox in row "
    print_for_playstation = "Value of playstation in row "

    var = {'xbox':xbox_count , 'playstation':playstation_count}
    sheet['E'+str(row)] = max(var,key=var.get)
    ps.save('data.xlsx')